from dataclasses import fields
import requests 
from bs4 import BeautifulSoup
import openpyxl

excel1 = openpyxl.Workbook()

print(excel1.sheetnames)

sheet = excel1.active

sheet.title = "Top Rated Movies"


print(excel1.sheetnames)


sheet.append(['Rank','Name','YOR','Rating'])

r =requests.get('https://www.imdb.com/chart/top/')

soup = BeautifulSoup(r.text ,'html.parser')

movies = soup.find('tbody',class_="lister-list").find_all('tr')

for movie in movies:
    rank = movie.find('td',class_='titleColumn').get_text(strip=True).split('.')
    ranked = rank[0]
    name = movie.find('td',class_="titleColumn").a.text
    published = movie.find('span',class_='secondaryInfo').text.strip('()')
    rating = movie.find('td',class_='ratingColumn imdbRating').strong.text
     
    print(ranked,name,published,rating)
    
    
    sheet.append([ranked,name,published,rating])
    
    
excel1.save("Top Rated Movies.xlsx")


